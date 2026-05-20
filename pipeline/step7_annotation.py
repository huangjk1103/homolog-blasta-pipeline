# coding = utf-8
"""
Step 7: 进化树注释 XLSX 生成

【输入】
    - input_fasta: str, FastTree 输入 FASTA（从序列 ID 提取基因组信息）
    - output_name: str, 输出文件名（不含扩展名）
    - output_dir: str, 输出目录
    - config: PipelineConfig, 配置参数

【输出】
    - {output_name}.xlsx: str, 树注释 Excel (Phylum/Class/Order/Family/Genus/Species)
    - .step7_annotation_done: str, 断点标记
"""

import os
import sys
import re
import logging
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from config import PipelineConfig, default_config

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def run_annotation(
    input_fasta: str,
    output_name: str,
    output_dir: str,
    config: PipelineConfig = None,
) -> str:
    """
    生成树注释 Excel

    Args:
        input_fasta: FastTree 输入 FASTA
        output_name: 输出文件名（不含 .xlsx）
        output_dir: 输出目录
        config: 配置对象

    Returns:
        XLSX 文件路径
    """
    if config is None:
        config = default_config

    os.makedirs(output_dir, exist_ok=True)

    # 检查断点
    checkpoint_file = os.path.join(output_dir, ".step7_annotation_done")
    xlsx_file = os.path.join(output_dir, f"{output_name}.xlsx")
    if os.path.exists(checkpoint_file) and os.path.exists(xlsx_file):
        logger.info("[STEP7] Resume: annotation xlsx already exists")
        return xlsx_file

    if not os.path.exists(input_fasta):
        raise FileNotFoundError(f"Input fasta not found: {input_fasta}")

    # 读取分类注释文件
    taxonomy_map = {}
    if os.path.exists(config.taxonomy_file):
        with open(config.taxonomy_file, "r") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split("\t")
                if len(parts) >= 7:
                    genome_id = parts[0].split(".")[0][-9:]  # 取最后9位
                    taxonomy_map[genome_id] = {
                        "Name": parts[0],
                        "Phylum": parts[1] if len(parts) > 1 else "",
                        "Class": parts[2] if len(parts) > 2 else "",
                        "Order": parts[3] if len(parts) > 3 else "",
                        "Family": parts[4] if len(parts) > 4 else "",
                        "Genus": parts[5] if len(parts) > 5 else "",
                        "Species": parts[6] if len(parts) > 6 else "",
                    }
        logger.info(f"[STEP7] Loaded {len(taxonomy_map)} taxonomy entries")

    # 从 FASTA 解析基因组 ID
    genome_ids = set()
    with open(input_fasta, "r") as f:
        for line in f:
            if line.startswith(">"):
                seq_id = line.strip().lstrip(">").split()[0]
                # 尝试提取基因组 ID（格式: GENOMEID_seqid）
                match = re.match(r"^(\S+?)_(\S+)", seq_id)
                if match:
                    gid = match.group(1)
                else:
                    gid = seq_id.split(".")[0][-9:]
                genome_ids.add(gid)

    logger.info(f"[STEP7] Found {len(genome_ids)} unique genome IDs in fasta")

    # 生成 XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TreeAnno"

    headers = ["Name", "Location", "Phylum", "Class", "Order", "Family", "Genus", "Species"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, gid in enumerate(sorted(genome_ids), 2):
        if gid in taxonomy_map:
            tax = taxonomy_map[gid]
            ws.cell(row=row_idx, column=1, value=tax["Name"])
            ws.cell(row=row_idx, column=2, value=tax.get("Location", ""))
            ws.cell(row=row_idx, column=3, value=tax["Phylum"])
            ws.cell(row=row_idx, column=4, value=tax["Class"])
            ws.cell(row=row_idx, column=5, value=tax["Order"])
            ws.cell(row=row_idx, column=6, value=tax["Family"])
            ws.cell(row=row_idx, column=7, value=tax["Genus"])
            ws.cell(row=row_idx, column=8, value=tax["Species"])
        else:
            ws.cell(row=row_idx, column=1, value=gid)
            for col in range(2, 9):
                ws.cell(row=row_idx, column=col, value="Unknown")

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    wb.save(xlsx_file)
    logger.info(f"[STEP7] Saved {len(genome_ids)} rows to {xlsx_file}")

    # 创建断点标记
    Path(checkpoint_file).touch()
    return xlsx_file


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python step7_annotation.py <input_fasta> <output_name> <output_dir>")
        sys.exit(1)

    run_annotation(sys.argv[1], sys.argv[2], sys.argv[3])
