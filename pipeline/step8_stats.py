# coding = utf-8
"""
Step 8: 基因计数统计

【输入】
    - input_fasta: str, 同源基因 FASTA 文件
    - output_dir: str, 输出目录
    - config: PipelineConfig, 配置参数

【输出】
    - homologs_statics.xlsx: str, 基因存在统计表
    - .step8_stats_done: str, 断点标记
"""

import os
import re
import logging
from pathlib import Path
from collections import Counter

from Bio import SeqIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def run_stats(
    input_fasta: str,
    output_dir: str,
    config,
) -> str:
    """
    生成基因存在/缺失统计表

    Args:
        input_fasta: 同源基因 FASTA 文件
        output_dir: 输出目录
        config: PipelineConfig 对象

    Returns:
        统计 xlsx 文件路径
    """
    os.makedirs(output_dir, exist_ok=True)

    # 检查断点
    checkpoint_file = os.path.join(output_dir, ".step8_stats_done")
    stat_xlsx = os.path.join(output_dir, "homologs_statics.xlsx")
    if os.path.exists(checkpoint_file) and os.path.exists(stat_xlsx):
        logger.info("[STEP8] Resume: statistics already exists")
        return stat_xlsx

    if not os.path.exists(input_fasta):
        raise FileNotFoundError(f"Input fasta not found: {input_fasta}")

    logger.info("[STEP8] Generating gene count statistics...")

    # 读取分类注释
    taxonomy_map = {}
    taxonomy_file = getattr(config, "taxonomy_file", None) or "/mnt/d/linux/1296_genome/homolog_blastp/phylogentic_tree_annotation.txt"
    if os.path.exists(taxonomy_file):
        with open(taxonomy_file, "r") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split("\t")
                if len(parts) >= 7:
                    gid = parts[0].split(".")[0][-9:]
                    taxonomy_map[gid] = {
                        "Name": parts[0],
                        "Phylum": parts[1] if len(parts) > 1 else "",
                        "Class": parts[2] if len(parts) > 2 else "",
                        "Order": parts[3] if len(parts) > 3 else "",
                        "Family": parts[4] if len(parts) > 4 else "",
                        "Genus": parts[5] if len(parts) > 5 else "",
                        "Species": parts[6] if len(parts) > 6 else "",
                    }

    # 解析序列 ID，统计每个基因组的基因数
    genome_counts = Counter()
    genome_sequences = {}

    for seq_record in SeqIO.parse(input_fasta, "fasta"):
        seq_id = seq_record.id.lstrip(">").split()[0]
        # 提取基因组 ID (格式: GENOMEID_seqid)
        match = re.match(r"^(\S+?)_(\S+)", seq_id)
        if match:
            gid = match.group(1)
        else:
            gid = seq_id.split(".")[0][-9:]

        genome_counts[gid] += 1
        if gid not in genome_sequences:
            genome_sequences[gid] = seq_record.id

    # 生成 xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gene Statistics"

    headers = ["Name", "GenomeID", "Phylum", "Class", "Order", "Family", "Genus", "Species", "Count"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, (gid, count) in enumerate(sorted(genome_counts.items(), key=lambda x: -x[1]), 2):
        tax = taxonomy_map.get(gid, {})
        ws.cell(row=row_idx, column=1, value=tax.get("Name", f"Genome_{gid}"))
        ws.cell(row=row_idx, column=2, value=gid)
        ws.cell(row=row_idx, column=3, value=tax.get("Phylum", "Unknown"))
        ws.cell(row=row_idx, column=4, value=tax.get("Class", ""))
        ws.cell(row=row_idx, column=5, value=tax.get("Order", ""))
        ws.cell(row=row_idx, column=6, value=tax.get("Family", ""))
        ws.cell(row=row_idx, column=7, value=tax.get("Genus", ""))
        ws.cell(row=row_idx, column=8, value=tax.get("Species", ""))
        ws.cell(row=row_idx, column=9, value=count)

    # 调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 2, 40)

    wb.save(stat_xlsx)
    logger.info(f"[STEP8] Statistics: {len(genome_counts)} genomes, {sum(genome_counts.values())} total hits")
    logger.info(f"[STEP8] Saved: {stat_xlsx}")

    # 创建断点标记
    Path(checkpoint_file).touch()
    return stat_xlsx


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python step8_stats.py <input_fasta> <output_dir> <taxonomy_file>")
        sys.exit(1)

    # 模拟 PipelineConfig 对象
    class FakeConfig:
        taxonomy_file = sys.argv[3]

    run_stats(sys.argv[1], sys.argv[2], FakeConfig())
